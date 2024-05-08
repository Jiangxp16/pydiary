from typing import OrderedDict

from core.util import encrypt_utils, sql_utils

sql_create = """
CREATE TABLE IF NOT EXISTS diary (
    id INTEGER PRIMARY KEY,
    content TEXT NOT NULL,
    weather CHAR(50) NOT NULL,
    location CHAR(50) NOT NULL
);
"""
sql_utils.cur.execute(sql_create)


class Diary:
    def __init__(self, id, content=None, weather=None, location=None):
        self.id = id
        self.content = content or ""
        self.weather = weather or ""
        self.location = location or ""

    def params(self):
        return (self.content, self.weather, self.location)

    def is_empty(self):
        return not (self.content or self.weather or self.location)


def add(diary: Diary = None, **kwargs):
    if diary is None:
        diary = Diary(**kwargs)
    if diary.is_empty():
        return False
    return sql_utils.execute("INSERT INTO diary (`id`, `content`, `weather`, `location`) VALUES (?,?,?,?)",
                             (diary.id, *diary.params()))


def get_by(**kwargs):
    sql_cmd = "SELECT `id`, `content`, `weather`, `location` FROM diary"
    args = []
    if len(kwargs) > 0:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    rs = sql_utils.select_one(sql_cmd, args)
    if rs is None:
        return None
    return Diary(*rs)


def get_between_dates(date1=0, date2=99999999) -> list[Diary]:
    rs_list = sql_utils.select("SELECT `id`, `content`, `weather`, `location` FROM diary WHERE id BETWEEN ? AND ? ORDER BY id",
                               (date1, date2))
    diaries = []
    for rs in rs_list:
        diaries.append(Diary(*rs))
    return diaries


def get_month_diary(year, month):
    date1 = year * 10000 + month * 100
    date2 = date1 + 31
    diaries = get_between_dates(date1, date2)
    return OrderedDict((diary.id, diary) for diary in diaries)


def update(diary: Diary):
    return sql_utils.execute("UPDATE diary SET content=?, weather=?, location=? WHERE id=?",
                             (*diary.params(), diary.id))


def update_many(diarys: list[Diary]):
    sql_cmd = "UPDATE diary SET content=?, weather=?, location=? WHERE id=?"
    return sql_utils.execute_many(sql_cmd, [(*diary.params(), diary.id) for diary in diarys])


def add_or_update(diary: Diary):
    if get_by(id=diary.id):
        return update(diary)
    return add(diary)


def delete(**kwargs):
    sql_cmd = "DELETE FROM diary"
    args = []
    if len(kwargs) > 0:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    return sql_utils.execute(sql_cmd, args)


def update_diary(diary: Diary):
    if diary.is_empty():
        return delete(id=diary.id)
    return add_or_update(diary)


def update_diaries(diaries: dict[int, Diary] | list[Diary]):
    res = True
    if isinstance(diaries, dict):
        diaries = list(diaries.values())
    diaries_to_update = []
    for diary in diaries:
        if diary.is_empty():
            res &= delete(id=diary.id)
        else:
            diaries_to_update.append(diary)
    res &= update_many(diaries_to_update)
    return res


def exp(file):
    diaries = get_between_dates()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(("Date", "Content", "Weather", "Location"))
    for diary in diaries:
        ws.append((diary.id, *diary.params()))
    wb.save(file)


def imp(file):
    diary_dict = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            diary = Diary(*(cell.value for cell in row))
            diary_dict[diary.id] = diary
        return update_diaries(diary_dict)
    except Exception as e:
        print(e.args)
        return False
