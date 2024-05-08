import datetime

from core.util import sql_utils
from core.util import utils

sql_create = """
CREATE TABLE IF NOT EXISTS note (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    begin INTEGER NOT NULL DEFAULT 0,
    last INTEGER NOT NULL DEFAULT 0,
    process INTEGER NOT NULL DEFAULT 0,
    desire INTEGER NOT NULL DEFAULT 0,
    priority INTEGER NOT NULL DEFAULT 0,
    content TEXT NOT NULL DEFAULT ""
);
"""
sql_utils.cur.execute(sql_create)


class Note:

    def __init__(self, id=None, begin=None, last=None, process=None, desire=None, priority=None, content=None):
        self.id = id
        self.begin = begin or utils.date2int(datetime.date.today())
        self.last = last or self.begin
        self.process = process or 0
        self.desire = desire or 0
        self.priority = priority or 0
        self.content = content or ""

    def params(self):
        return (self.begin, self.last, self.process, self.desire, self.priority, self.content)

    def __str__(self):
        return str(self.params())


def get_last():
    rs = sql_utils.get_last("note")
    if rs is not None:
        return Note(*rs)
    return None


def add(note=None, **kwargs):
    if note is None:
        note = Note(**kwargs)
    if sql_utils.execute("INSERT INTO note (`begin`, `last`, `process`, `desire`, `priority`, `content`) VALUES (?,?,?,?,?,?)",
                         note.params()):
        return get_last()
    return None


def add_many(notes):
    sql_cmd = "INSERT INTO note (`begin`, `last`, `process`, `desire`, `priority`, `content`) VALUES (?,?,?,?,?,?)"
    data_list = [note.params() for note in notes]
    return sql_utils.execute_many(sql_cmd, data_list)


def get_list_by(**kwargs):
    sql_cmd = "SELECT `id`, `begin`, `last`, `process`, `desire`, `priority`, `content` FROM note"
    args = []
    if kwargs:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    rs_list = sql_utils.select(sql_cmd, args)
    return [Note(*rs) for rs in rs_list]


def update(note: Note):
    sql_cmd = "UPDATE note SET `begin`=?, `last`=?, `process`=?, `desire`=?, \
        `priority`=?, `content`=? WHERE `id`=?"
    return sql_utils.execute(sql_cmd, (*note.params(), note.id))


def update_many(notes: list[Note]):
    sql_cmd = "UPDATE note SET `begin`=?, `last`=?, `process`=?, `desire`=?, \
        `priority`=?, `content`=? WHERE `id`=?"
    return sql_utils.execute_many(sql_cmd, [(*note.params(), note.id) for note in notes])


def delete(**kwargs):
    sql_cmd = "DELETE FROM note"
    args = []
    if len(kwargs) > 0:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    return sql_utils.execute(sql_cmd, args)


def exp(file):
    notes = get_list_by()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(("ID", "begin", "last", "process", "desire", "priority", "content"))
    for note in notes:
        ws.append((note.id, *note.params()))
    wb.save(file)


def imp(file):
    notes = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            notes.append(Note(*(cell.value for cell in row)))
    except Exception as e:
        print(e.args)
    return add_many(notes)
