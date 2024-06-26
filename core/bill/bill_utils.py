import datetime

from core.util import sql_utils
from core.util import utils

sql_create = """
CREATE TABLE IF NOT EXISTS bill (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date INTEGER NOT NULL DEFAULT 0,
    inout INTEGER NOT NULL DEFAULT -1,
    type CHAR(20) NOT NULL DEFAULT "",
    amount DECIMAL(10,2) NOT NULL DEFAULT 0.0,
    item TEXT NOT NULL DEFAULT ""
);
"""
sql_utils.cur.execute(sql_create)

# sql_create_index = "CREATE INDEX IF NOT EXISTS index_date ON bill (date);"
# sqlutils.cur.execute(sql_create_index)
# sqlutils.conn.commit()


class Bill:

    def __init__(self, id=None, date=None, inout=None, type=None, amount=None, item=None):
        self.id = id
        self.date = date or utils.date2int(datetime.date.today())
        self.inout = inout or -1
        self.type = type or ""
        self.amount = amount or 0.0
        self.item = item or ""

    def params(self):
        return (self.date, self.inout, self.type, self.amount, self.item)

    def __str__(self) -> str:
        return str(self.params())


def get_last():
    rs = sql_utils.get_last("bill")
    if rs is not None:
        return Bill(*rs)
    return None


def add(bill=None, **kwargs):
    if bill is None:
        bill = Bill(**kwargs)
    if sql_utils.execute("INSERT INTO bill (`date`, `inout`, `type`, `amount`, `item`) VALUES (?,?,?,?,?)",
                         bill.params()):
        return get_last()
    return None


def add_many(bills):
    sql_cmd = "INSERT INTO bill (`date`, `inout`, `type`, `amount`, `item`) VALUES (?,?,?,?,?)"
    data_list = [bill.params() for bill in bills]
    return sql_utils.execute_many(sql_cmd, data_list)


def get_list_by(**kwargs):
    sql_cmd = "SELECT `id`, `date`, `inout`, `type`, `amount`, `item` FROM bill"
    args = []
    if kwargs:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    rs_list = sql_utils.select(sql_cmd, args)
    return [Bill(*rs) for rs in rs_list]


def get_between_dates(date1=0, date2=99999999):
    sql_cmd = "SELECT `id`, `date`, `inout`, `type`, `amount`, `item` FROM bill WHERE `date` BETWEEN ? AND ? ORDER BY `date` ASC"
    rs_list = sql_utils.select(sql_cmd, (date1, date2))
    return [Bill(*rs) for rs in rs_list]


def update(bill: Bill):
    bill.item = bill.item.strip()
    sql_cmd = "UPDATE bill SET `date`=?, `inout`=?, `type`=?, `amount`=?, `item`=? WHERE `id`=?"
    return sql_utils.execute(sql_cmd, (*bill.params(), bill.id))


def update_many(bills: list[Bill]):
    sql_cmd = "UPDATE bill SET `date`=?, `inout`=?, `type`=?, `amount`=?, `item`=? WHERE `id`=?"
    for bill in bills:
        bill.item = bill.item.strip()
    return sql_utils.execute_many(sql_cmd, [(*bill.params(), bill.id) for bill in bills])


def delete(**kwargs):
    sql_cmd = "DELETE FROM bill"
    args = []
    if len(kwargs) > 0:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    return sql_utils.execute(sql_cmd, args)


def exp(file, date1=0, date2=99999999):
    bills = get_between_dates(date1, date2)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(("ID", "Date", "Inout", "Type", "Amount", "Remark"))
    for bill in bills:
        ws.append((bill.id, *bill.params()))
    wb.save(file)


def imp(file):
    bills = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            bills.append(Bill(*(cell.value for cell in row)))
    except Exception as e:
        print(e.args)
    return add_many(bills)
