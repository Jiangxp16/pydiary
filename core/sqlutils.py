import datetime
import sqlite3
from typing import OrderedDict
from core import utils
db_name = utils.load_config("global", "db_name")
db_path = utils.get_path(db_name)
conn = sqlite3.connect(db_name)
cur = conn.cursor()
sql_create = """
CREATE TABLE IF NOT EXISTS diary (
    id INTEGER PRIMARY KEY,
    content TEXT NOT NULL,
    weather CHAR(50) NOT NULL,
    location CHAR(50) NOT NULL
);
"""
cur.execute(sql_create)


def close_connection():
    try:
        cur.close()
        conn.close()
    except Exception as e:
        print(e.args)


def date2int(date):
    return date.year * 10000 + date.month * 100 + date.day


def int2date(x):
    return datetime.date(int(x / 10000), int((x % 10000) / 100), int(x % 100))


def add(date, content, weather='', location=''):
    if not content and not weather and not location:
        return
    _id = date2int(date) if isinstance(date, datetime.date) else date
    try:
        cur.execute("INSERT INTO diary (id, content, weather, location) VALUES (?,?,?,?)",
                    (_id, content, weather, location))
        conn.commit()
    except Exception as e:
        print(e.args)
        conn.rollback()
        return False
    return True


def get_by_date(date: datetime.date):
    _id = date2int(date) if isinstance(date, datetime.date) else date
    cur.execute("SELECT id, content, weather, location FROM diary WHERE id=?", (_id,))
    return cur.fetchone()


def get_between_dates(date1: datetime.date, date2: datetime.date):
    _id1 = date2int(date1) if isinstance(date1, datetime.date) else date1
    _id2 = date2int(date2) if isinstance(date2, datetime.date) else date2
    cur.execute("SELECT id, content, weather, location FROM diary WHERE id BETWEEN ? AND ? ORDER BY id", (_id1, _id2))
    return cur.fetchall()


def get_diary_dict(date1: datetime.date, date2: datetime.date):
    diaries = get_between_dates(date1, date2)
    diary_dict = OrderedDict()
    for diary in diaries:
        diary_dict[diary[0]] = list(diary[1:4])
    return diary_dict


def get_month_diary(year, month):
    _id1 = year * 10000 + month * 100
    _id2 = _id1 + 31
    diaries = cur.execute("SELECT id, content, weather, location FROM diary WHERE id BETWEEN ? AND ? ORDER BY id", (_id1, _id2))
    diary_dict = OrderedDict()
    for diary in diaries:
        diary_dict[diary[0]] = list(diary[1:4])
    return diary_dict


def update(date, content, weather='', location=''):
    _id = date2int(date) if isinstance(date, datetime.date) else date
    try:
        cur.execute("UPDATE diary set content=?, weather=?, location=? where id=?", (content, weather, location, _id))
        conn.commit()
    except Exception as e:
        print(e.args)
        conn.rollback()
        return False
    return True


def add_or_update(date, content, weather='', location=''):
    if get_by_date(date):
        return update(date, content, weather, location)
    return add(date, content, weather, location)


def delete(date: datetime.date | int):
    _id = date2int(date) if isinstance(date, datetime.date) else date
    try:
        cur.execute("DELETE FROM diary where id=?", (_id,))
        conn.commit()
    except Exception as e:
        print(e.args)
        conn.rollback()
        return False
    return True


def update_diaries(diary_dict):
    res = True
    for _id in diary_dict:
        if not diary_dict[_id][0] and not diary_dict[_id][1] and not diary_dict[_id][2]:
            res &= delete(_id)
        else:
            res &= add_or_update(_id, diary_dict[_id][0], diary_dict[_id][1], diary_dict[_id][2])
    return res


def update_diary(date, diary):
    if not diary[0] and not diary[1] and not diary[2]:
        return delete(date)
    return add_or_update(date, diary[0], diary[1], diary[2])


def export_to_file(file):
    diaries = get_between_dates(0, 99999999)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(("Date", "Content", "Weather", "Location"))
    for diary in diaries:
        ws.append(list(diary))
    wb.save(file)


def import_from_file(file):
    diary_dict = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            diary = [cell.value for cell in row]
            diary_dict[diary[0]] = diary[1:4]
        return update_diaries(diary_dict)
    except Exception as e:
        print(e.args)
        return False
