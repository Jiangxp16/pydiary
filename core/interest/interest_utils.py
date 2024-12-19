import datetime

from core.util import sql_utils, utils

sql_create = """
CREATE TABLE IF NOT EXISTS interest (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    added INTEGER NOT NULL DEFAULT 0,
    updated INTEGER NOT NULL DEFAULT 0,
    name CHAR(255) NOT NULL DEFAULT '',
    sort INTEGER NOT NULL DEFAULT 0,
    progress CHAR(20) NOT NULL DEFAULT '',
    publish INTEGER NOT NULL DEFAULT 0,
    date INTEGER NOT NULL DEFAULT 0,
    score_db DECIMAL(2,1) NOT NULL DEFAULT -1,
    score_imdb DECIMAL(2,1) NOT NULL DEFAULT -1,
    score DECIMAL(2,1) NOT NULL DEFAULT -1,
    remark TEXT NOT NULL
);
"""
sql_utils.cur.execute(sql_create)

# sql_create_index = "CREATE INDEX IF NOT EXISTS index_sort ON interest (sort);"
# sqlutils.cur.execute(sql_create_index)
# sqlutils.conn.commit()

sql_update = "UPDATE interest SET `sort`=7 WHERE `sort`=0"
sql_utils.cur.execute(sql_update)
sql_utils.conn.commit()


class Interest:

    fields = ["added", "updated", "name", "sort", "progress", "publish",
              "date", "score_db", "score_imdb", "score", "remark"]

    def __init__(self, id=None, added=None, updated=None, name=None, sort=None, progress=None, publish=None,
                 date=None, score_db=None, score_imdb=None, score=None, remark=None):
        self.id = id
        self.added = added or utils.date2int(datetime.date.today())
        self.updated = updated or 0
        self.name = name or ""
        self.sort = sort or 7
        self.progress = progress or ""
        self.publish = publish or 0
        self.date = date or 0
        self.score_db = score_db or -1
        self.score_imdb = score_imdb or -1
        self.score = score or -1
        self.remark = remark or ""

    def params(self):
        return (self.added, self.updated, self.name, self.sort, self.progress, self.publish,
                self.date, self.score_db, self.score_imdb, self.score, self.remark)

    def set_param(self, field, value):
        if isinstance(field, str) and hasattr(self, field):
            setattr(self, field, value)
            return True
        elif isinstance(field, int):
            if field >= 0 and field <= len(self.fields):
                return self.set_param(self.fields[field], value)
        return False

    def get_param(self, field):
        if isinstance(field, str) and hasattr(self, field):
            return getattr(self, field)
        elif isinstance(field, int):
            if field >= 0 and field <= len(self.fields):
                return self.get_param(self.fields[field])
        return None

    def __str__(self):
        return str(self.params())


def get_last():
    rs = sql_utils.get_last("interest")
    if rs is not None:
        return Interest(*rs)
    return None


def add(interest=None, **kwargs):
    if interest is None:
        interest = Interest(**kwargs)
        interest.updated = utils.date2int(datetime.date.today())
        interest.added = utils.date2int(datetime.date.today())
        interest.date = utils.date2int(datetime.date.today()) // 100
    if sql_utils.execute("INSERT INTO interest (`added`, `updated`, `name`, `sort`, `progress`, `publish`, `date`, \
                           `score_db`, `score_imdb`, `score`, `remark`) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                         interest.params()):
        return get_last()
    return None


def add_many(interests: list[Interest]):
    sql_cmd = "INSERT INTO interest (`added`, `updated`, `name`, `sort`, `progress`, `publish`, `date`, \
                           `score_db`, `score_imdb`, `score`, `remark`) VALUES (?,?,?,?,?,?,?,?,?,?,?)"
    data_list = [interest.params() for interest in interests]
    return sql_utils.execute_many(sql_cmd, data_list)


def get_list_by(**kwargs):
    if "sort" in kwargs and kwargs["sort"] == 0:
        kwargs.pop("sort")
    interests = []
    sql_cmd = "SELECT `id`, `added`, `updated`, `name`, `sort`, `progress`, `publish`, `date`, \
                           `score_db`, `score_imdb`, `score`, `remark` FROM interest"
    args = []
    if len(kwargs) > 0:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    sql_cmd += " ORDER BY `id` ASC"
    rs_list = sql_utils.select(sql_cmd, args)
    for rs in rs_list:
        interests.append(Interest(*rs))
    return interests


def update(interest: Interest):
    interest.updated = utils.date2int(datetime.date.today())
    interest.name = interest.name.strip()
    interest.progress = interest.progress.strip()
    interest.remark = interest.remark.strip()
    interest.sort = interest.sort or 7
    sql_cmd = "UPDATE interest SET `added`=?, `updated`=?, `name`=?, `sort`=?, `progress`=?, `publish`=?, `date`=?,\
                           `score_db`=?, `score_imdb`=?, `score`=?, `remark`=? WHERE `id`=?"
    return sql_utils.execute(sql_cmd, (*interest.params(), interest.id))


def update_many(interests: list[Interest]):
    today = utils.date2int(datetime.date.today())
    for interest in interests:
        interest.updated = today
        interest.name = interest.name.strip()
        interest.progress = interest.progress.strip()
        interest.remark = interest.remark.strip()
    sql_cmd = "UPDATE interest SET `added`=?, `updated`=?, `name`=?, `sort`=?, `progress`=?, `publish`=?, `date`=?,\
                           `score_db`=?, `score_imdb`=?, `score`=?, `remark`=? WHERE `id`=?"
    data_list = [(*interest.params(), interest.id) for interest in interests]
    return sql_utils.execute_many(sql_cmd, data_list)


def add_or_update(interest):
    if interest.id is not None and get_list_by(id=interest.id):
        return update(interest)
    return add(interest) is not None


def delete(**kwargs):
    if "sort" in kwargs and kwargs["sort"] == 0:
        kwargs.pop("sort")
    if len(kwargs) == 0:
        try:
            sql_utils.cur.execute("DROP TABLE IF EXISTS interest;")
            sql_utils.conn.commit()
            sql_utils.cur.execute(sql_create)
            return True
        except Exception as e:
            print(e.args)
            return False
    sql_cmd = "DELETE FROM interest"
    args = []
    if len(kwargs) > 0:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    return sql_utils.execute(sql_cmd, args)


def exp(file, sort=0):
    interests = get_list_by(sort=sort)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(("id", "Added", "Updated", "Name", "Sort", "Progress",
               "Publish", "Date", "Score_db", "Score_imdb", "Score", "Remark"))
    for interest in interests:
        ws.append((interest.id, *interest.params()))
    wb.save(file)


def imp(file):
    interests = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            interest = Interest(*(cell.value for cell in row))
            interests.append(interest)
        return add_many(interests)
    except Exception as e:
        print(e.args)
        return False
