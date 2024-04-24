import datetime
from typing import OrderedDict
from core import sqlutils, utils

sql_create = """
CREATE TABLE IF NOT EXISTS diary (
    id INTEGER PRIMARY KEY,
    content TEXT NOT NULL,
    weather CHAR(50) NOT NULL,
    location CHAR(50) NOT NULL
);
"""
sqlutils.cur.execute(sql_create)


class Diary:
    def __init__(self, id, content=None, weather=None, location=None):
        self.id = id
        self.content = content or ""
        self.weather = weather or ""
        self.location = location or ""

    def params(self):
        return (self.content, self.weather, self.location)

    def is_empty(self):
        return not (self.content or self.weather or self.location)


def add(diary: Diary = None, **kwargs):
    if diary is None:
        diary = Diary(**kwargs)
    if diary.is_empty():
        return False
    return sqlutils.insert("INSERT INTO diary (`id`, `content`, `weather`, `location`) VALUES (?,?,?,?)",
                           (diary.id, *diary.params()))


def get_by(**kwargs):
    sql_cmd = "SELECT `id`, `content`, `weather`, `location` FROM diary"
    args = []
    if len(kwargs) > 0:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    rs = sqlutils.select_one(sql_cmd, args)
    if rs is None:
        return None
    return Diary(*rs)


def get_between_dates(date1, date2) -> list[Diary]:
    rs_list = sqlutils.select("SELECT `id`, `content`, `weather`, `location` FROM diary WHERE id BETWEEN ? AND ? ORDER BY id",
                              (date1, date2))
    diaries = []
    for rs in rs_list:
        diaries.append(Diary(*rs))
    return diaries


def get_month_diary(year, month):
    date1 = year * 10000 + month * 100
    date2 = date1 + 31
    diaries = get_between_dates(date1, date2)
    return OrderedDict((diary.id, diary) for diary in diaries)


def update(diary: Diary):
    return sqlutils.update("UPDATE diary SET content=?, weather=?, location=? WHERE id=?",
                           (*diary.params(), diary.id))


def add_or_update(diary: Diary):
    if get_by(id=diary.id):
        return update(diary)
    return add(diary)


def delete(**kwargs):
    sql_cmd = "DELETE FROM diary"
    args = []
    if len(kwargs) > 0:
        sql_cmd += " WHERE "
        for key in kwargs:
            sql_cmd += "`%s`=? AND " % key
            args.append(kwargs[key])
        sql_cmd = sql_cmd[:-5]
    return sqlutils.delete(sql_cmd, args)


def update_diary(diary: Diary):
    if diary.is_empty():
        return delete(id=diary.id)
    return add_or_update(diary)


def update_diaries(diary_dict: dict[int, Diary]):
    res = True
    for diary in diary_dict.values():
        res &= update_diary(diary)
    return res


def exp(file):
    diaries = get_between_dates(0, 99999999)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(("Date", "Content", "Weather", "Location"))
    for diary in diaries:
        ws.append([diary.id, *diary.params()])
    wb.save(file)


def imp(file):
    diary_dict = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            diary = Diary(*(cell.value for cell in row))
            diary_dict[diary.id] = diary
        return update_diaries(diary_dict)
    except Exception as e:
        print(e.args)
        return False
